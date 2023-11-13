from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string


def sendemail(firstname='hossein', lastname='yazdi', email='iran581@gmail.com'):
    subject, from_email = 'hello', 'info.xtrader@gmail.com'
    text_content = 'This is an important message.'
    html_content = render_to_string('invetemail.html', {'firstname': firstname, 'lastname': lastname})
    msg = EmailMultiAlternatives(subject, text_content, from_email, [email])
    msg.attach_alternative(html_content, "text/html")
    msg.send()
#

import openpyxl

wb = openpyxl.load_workbook('Users.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
i = 1
while (sheet.cell(row=i, column=12).value != None):
    print('Email: ' + sheet.cell(row=i, column=12).value, end='')
    if (sheet.cell(row=i, column=13).value != None):
        print(' Name: ' + sheet.cell(row=i, column=13).value, end='')
    if (sheet.cell(row=i, column=14).value != None):
        print(' Name: ' + sheet.cell(row=i, column=14).value)
    i += 1
